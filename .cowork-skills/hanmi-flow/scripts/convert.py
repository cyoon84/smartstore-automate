#!/usr/bin/env python3
"""
스마트스토어 출고 엑셀 → 한미택배 업로드용 엑셀 변환.

사용법:
    python3 convert.py <스마트스토어_경로> <비밀번호>

비밀번호가 없으면 빈 문자열로 호출.
"""
import json
import io
import os
import sys
from datetime import datetime

import msoffcrypto
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def main():
    if len(sys.argv) < 2:
        print("Usage: convert.py <smartstore_path> [password]", file=sys.stderr)
        sys.exit(1)

    smartstore_file = sys.argv[1]
    password = sys.argv[2] if len(sys.argv) > 2 else ""

    hanmi_template = os.path.expanduser("~/smartstore-project/templates/hanmi-form.xls")
    product_mapping = os.path.expanduser("~/smartstore-project/templates/product-mapping.xlsx")
    config_file = os.path.expanduser("~/smartstore-project/config.json")

    with open(config_file) as f:
        cfg = json.load(f)
    business_id = cfg["business_id"]
    sender_name = cfg["sender_name"]
    sender_email = cfg["sender_email"]
    sender_phone = cfg["sender_phone"]
    sender_address = cfg["sender_address"]

    # 스마트스토어 파일 복호화 (비밀번호 없으면 그대로 읽기)
    decrypted = io.BytesIO()
    with open(os.path.expanduser(smartstore_file), "rb") as f:
        if password:
            office_file = msoffcrypto.OfficeFile(f)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
            decrypted.seek(0)
        else:
            decrypted = io.BytesIO(f.read())
    ss = pd.read_excel(decrypted, engine="openpyxl", header=1)

    # 한미택배 컬럼 읽기
    hanmi_df = pd.read_excel(hanmi_template, header=0)
    hanmi_cols = [col.split(".")[0] for col in hanmi_df.columns]

    # 상품 매핑
    mapping_df = pd.read_excel(product_mapping, dtype={"상품번호": str})
    mapping_df["상품번호"] = mapping_df["상품번호"].str.split(".").str[0]
    mapping_df = mapping_df.drop_duplicates(subset="상품번호", keep="first")
    product_map = mapping_df.set_index("상품번호").to_dict("index")

    def val(row, col):
        v = row.get(col)
        return "" if pd.isna(v) else v

    def phone(row, col):
        v = val(row, col)
        return str(v).replace("-", "") if v != "" else ""

    def zipcode(row, col):
        v = val(row, col)
        if v == "":
            return ""
        s = str(v).replace(".0", "").strip()
        return s.zfill(5) if s.isdigit() else s

    def mget(pmap, col):
        if pmap is None:
            return ""
        v = pmap.get(col, "")
        return "" if pd.isna(v) else (str(v) if v != "" else "")

    # 같은 주문 + 같은 상품번호 → 수량 합산
    ss["_product_num"] = ss["상품번호"].astype(str).str.split(".").str[0]
    ss["_order_num"] = ss["주문번호"].astype(str)
    agg_dict = {col: "first" for col in ss.columns if col not in ["수량", "_product_num", "_order_num"]}
    agg_dict["수량"] = "sum"
    ss = ss.groupby(["_order_num", "_product_num"], sort=False).agg(agg_dict).reset_index()

    hs_col_idx = hanmi_cols.index("HS CODE")
    zipcode_col_idx = next(i for i, c in enumerate(hanmi_cols) if c == "우편번호")

    missing_products = []
    rows = []
    seen_orders = {}
    row_num = 1

    for _, row in ss.iterrows():
        product_num = row["_product_num"]
        order_num = row["_order_num"]
        pmap = product_map.get(product_num)

        eng_name = mget(pmap, "상품명(영문)")
        hs_code = mget(pmap, "HS CODE")
        brand = mget(pmap, "브랜드")
        unit_price = mget(pmap, "단가")
        site_url = mget(pmap, "SITE URL")
        seller = mget(pmap, "해외판매자 상호")

        if not eng_name:
            missing_products.append((product_num, val(row, "상품명")))

        is_first = order_num not in seen_orders
        if is_first:
            seen_orders[order_num] = row_num

        new_row = [""] * len(hanmi_cols)
        new_row[hs_col_idx] = hs_code
        new_row[hs_col_idx + 1] = ""
        new_row[hs_col_idx + 2] = eng_name if eng_name else val(row, "상품명")
        new_row[hs_col_idx + 3] = brand
        new_row[hs_col_idx + 4] = unit_price if unit_price else val(row, "상품가격")
        new_row[hs_col_idx + 5] = val(row, "수량")
        new_row[hs_col_idx + 6] = site_url
        new_row[hs_col_idx + 7] = ""
        new_row[hs_col_idx + 8] = "B"
        new_row[hs_col_idx + 9] = ""
        new_row[hs_col_idx + 10] = seller
        new_row[hs_col_idx + 11] = ""
        new_row[hs_col_idx + 12] = sender_name
        new_row[hs_col_idx + 13] = ""
        new_row[hs_col_idx + 14] = ""
        new_row[hs_col_idx + 15] = val(row, "주문번호")

        if is_first:
            new_row[0] = row_num
            new_row[1] = business_id
            new_row[2] = sender_name
            new_row[3] = sender_email
            new_row[4] = sender_phone
            new_row[5] = sender_address
            new_row[6] = 1
            new_row[7] = val(row, "수취인명")
            new_row[8] = phone(row, "수취인연락처1")
            new_row[9] = phone(row, "수취인연락처2")
            new_row[10] = zipcode(row, "우편번호")
            new_row[11] = val(row, "기본배송지")
            new_row[12] = val(row, "상세배송지")
            new_row[13] = val(row, "개인통관고유부호")
            new_row[14] = val(row, "배송메세지")
            new_row[15] = 1
            new_row[16] = "a"
            new_row[17] = 1
            new_row[18] = 1
            new_row[19] = 1
            new_row[20] = 1
            new_row[21] = 1
            new_row[22] = 1
            row_num += 1

        rows.append(new_row)

    if missing_products:
        lines = [f"  상품번호 {num}: {name}" for num, name in missing_products]
        print("MISSING:" + "\n".join(lines))

    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill("solid", start_color="366092", end_color="366092")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    data_font = Font(name="Arial", size=9)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(hanmi_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    zipcode_col = zipcode_col_idx + 1

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, v in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.font = data_font
            cell.border = border
            cell.alignment = (
                center_align if col_idx in [1, 6, 15, 16, 17, 18, 19, 20, 21, 22, 28, 31] else left_align
            )
            if col_idx == zipcode_col:
                cell.number_format = "@"

    col_widths = {
        1: 6, 2: 14, 3: 14, 4: 20, 5: 13, 6: 50, 7: 10, 8: 16, 9: 14, 10: 14,
        11: 10, 12: 40, 13: 20, 14: 16, 15: 20, 16: 12, 17: 50, 18: 10, 19: 10,
        20: 10, 21: 8, 22: 10, 23: 8, 24: 12, 25: 10, 26: 45, 27: 12, 28: 10,
        29: 8, 30: 15, 31: 15, 32: 45, 33: 15, 34: 15, 35: 15, 36: 15, 37: 15,
        38: 15, 39: 40,
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 40

    today = datetime.now().strftime("%Y%m%d")
    output_dir = os.path.expanduser("~/smartstore-project/output/한미택배")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"한미택배_업로드용_{today}.xlsx")
    wb.save(output_path)
    print(f"✅ 변환 완료: {output_path} ({len(rows)}건)")


if __name__ == "__main__":
    main()
