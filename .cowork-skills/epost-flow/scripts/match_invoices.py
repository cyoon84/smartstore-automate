#!/usr/bin/env python3
"""
우체국택배 송장 엑셀과 스마트스토어 출고 파일을 매칭해 발송처리용 엑셀을 생성.

사용법:
    python3 match_invoices.py <스마트스토어_경로> <비밀번호> <우체국송장_경로>

비밀번호가 없으면 빈 문자열로 호출. 송장번호 컬럼에 우체국 `등기번호` 값을 입력한다.
"""
import io
import os
import sys
from datetime import datetime

import msoffcrypto
import pandas as pd
from openpyxl import load_workbook


def normalize_phone(v):
    if pd.isna(v):
        return ""
    s = str(v).replace("-", "").replace(" ", "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.lstrip("0")


def main():
    if len(sys.argv) < 4:
        print(
            "Usage: match_invoices.py <smartstore_path> <password> <epost_invoice_path>",
            file=sys.stderr,
        )
        sys.exit(1)

    smartstore_file = sys.argv[1]
    password = sys.argv[2]
    epost_invoice_file = sys.argv[3]

    # 우체국 송장 — 받는분+전화번호로 매핑
    epost = pd.read_excel(os.path.expanduser(epost_invoice_file), header=0)
    name_phone_to_tracking = {}
    for _, r in epost.iterrows():
        key = (str(r["받는분"]).strip(), normalize_phone(r["받는분 전화번호"]))
        name_phone_to_tracking[key] = str(r["등기번호"])

    # 스마트스토어 복호화
    decrypted = io.BytesIO()
    with open(os.path.expanduser(smartstore_file), "rb") as f:
        if password:
            office_file = msoffcrypto.OfficeFile(f)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
            decrypted.seek(0)
        else:
            decrypted = io.BytesIO(f.read())
    wb = load_workbook(decrypted)
    ws = wb.active

    # 2행에서 컬럼 인덱스 찾기
    col_index = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(2, col).value
        if v:
            col_index[v] = col

    song_jang_col = col_index.get("송장번호")
    name_col = col_index.get("수취인명")
    phone_col = col_index.get("수취인연락처1")

    filled = 0
    unmatched = []
    matched_names = []  # 매칭된 수취인 이름 (중복 제거, 첫 등장 순서)
    seen_matched = set()
    for row in range(3, ws.max_row + 1):
        name = ws.cell(row, name_col).value
        if not name:
            continue
        phone = normalize_phone(ws.cell(row, phone_col).value)
        key = (str(name).strip(), phone)
        tracking = name_phone_to_tracking.get(key)
        if tracking:
            ws.cell(row, song_jang_col).value = tracking
            filled += 1
            nm = str(name).strip()
            if nm not in seen_matched:
                matched_names.append(nm)
                seen_matched.add(nm)
        else:
            unmatched.append((str(name).strip(), phone))

    ws.delete_rows(1)
    ws.title = "발송처리"

    today = datetime.now().strftime("%Y%m%d")
    output_dir = os.path.expanduser("~/smartstore-project/output/발송처리")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"스마트스토어_발송처리_{today}-우체국.xlsx")
    wb.save(output_path)

    print(f"DONE:{output_path}:{filled}")
    if matched_names:
        print("RECIPIENTS:" + "|".join(matched_names))
    if unmatched:
        print("UNMATCHED:" + "|".join([f"{n}({p})" for n, p in unmatched]))


if __name__ == "__main__":
    main()
